"""
Step 15: Create an Excel validation workbook.

Output created:
- outputs/market_risk_validation.xlsx

What it does:
- exports small samples from the SQLite database
- adds Excel formulas that independently recalculate key metrics
- helps prove your Python/SQL calculations are correct

Note:
This workbook is for QA. It is not the final dashboard.
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import DB_PATH, OUTPUT_DIR, TABLES
from db_utils import connect, read_sql, require_tables

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")


def _style_sheet(ws) -> None:
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(max_len + 2, 12), 28)


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    _style_sheet(ws)


def _create_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    rows = [
        ["Validation Workbook", "Market Risk & Portfolio Exposure Analytics Platform"],
        ["Purpose", "Spot-check Python/SQL calculations in Excel before presenting the dashboard."],
        ["Return_Check", "Recalculates daily return as Adj Close / Prior Adj Close - 1."],
        ["Drawdown_Check", "Recalculates running peak and drawdown."],
        ["Portfolio_Check", "Recalculates portfolio return as sum(weight x stock return)."],
        ["Benchmark_Check", "Compares portfolio vs benchmark returns and active return."],
        ["How to use", "Open each sheet and confirm diff columns are approximately zero."],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    ws["B1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90


def _return_check(conn) -> pd.DataFrame:
    query = f"""
        SELECT
            s.ticker,
            p.date,
            p.adj_close,
            r.daily_return AS python_sql_daily_return
        FROM {TABLES['price']} p
        JOIN {TABLES['stock']} s ON p.stock_id = s.stock_id
        LEFT JOIN {TABLES['return']} r
            ON p.stock_id = r.stock_id AND p.date = r.date
        WHERE s.stock_id = (
            SELECT MIN(stock_id)
            FROM {TABLES['stock']}
            WHERE UPPER(ticker) <> 'SPY'
        )
        ORDER BY p.date
        LIMIT 15;
    """
    return read_sql(conn, query)


def _drawdown_check(conn) -> pd.DataFrame:
    query = f"""
        SELECT
            s.ticker,
            p.date,
            p.adj_close,
            rm.drawdown AS python_sql_drawdown
        FROM {TABLES['price']} p
        JOIN {TABLES['stock']} s ON p.stock_id = s.stock_id
        LEFT JOIN {TABLES['risk_metric']} rm
            ON p.stock_id = rm.stock_id AND p.date = rm.date
        WHERE s.stock_id = (
            SELECT MIN(stock_id)
            FROM {TABLES['stock']}
            WHERE UPPER(ticker) <> 'SPY'
        )
        ORDER BY p.date
        LIMIT 30;
    """
    return read_sql(conn, query)


def _portfolio_check(conn) -> pd.DataFrame:
    query = f"""
        WITH normalized_holdings AS (
            SELECT stock_id, CASE WHEN weight > 1 THEN weight / 100.0 ELSE weight END AS weight
            FROM {TABLES['portfolio_holding']}
        ), latest_dates AS (
            SELECT date
            FROM {TABLES['return']}
            GROUP BY date
            ORDER BY date DESC
            LIMIT 10
        )
        SELECT
            r.date,
            s.ticker,
            nh.weight,
            r.daily_return,
            nh.weight * r.daily_return AS python_sql_contribution
        FROM latest_dates ld
        JOIN {TABLES['return']} r ON ld.date = r.date
        JOIN normalized_holdings nh ON r.stock_id = nh.stock_id
        JOIN {TABLES['stock']} s ON r.stock_id = s.stock_id
        ORDER BY r.date, s.ticker;
    """
    return read_sql(conn, query)


def _benchmark_check(conn) -> pd.DataFrame:
    query = f"""
        SELECT
            date,
            portfolio_return,
            benchmark_return,
            active_return,
            portfolio_value,
            benchmark_value
        FROM {TABLES['portfolio_benchmark_comparison']}
        ORDER BY date DESC
        LIMIT 15;
    """
    return read_sql(conn, query)


def create_excel_validation_workbook() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / "market_risk_validation.xlsx"

    with connect(DB_PATH) as conn:
        require_tables(
            conn,
            [
                TABLES["stock"],
                TABLES["price"],
                TABLES["return"],
                TABLES["risk_metric"],
                TABLES["portfolio_holding"],
                TABLES["portfolio_benchmark_comparison"],
            ],
        )

        wb = Workbook()
        _create_readme(wb)

        # Return check sheet.
        return_df = _return_check(conn)
        ws = wb.create_sheet("Return_Check")
        _write_dataframe(ws, return_df)
        ws.cell(row=1, column=5, value="excel_return_formula")
        ws.cell(row=1, column=6, value="diff_vs_python_sql")
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=5, value=f"=C{row}/C{row-1}-1")
            ws.cell(row=row, column=6, value=f"=E{row}-D{row}")
        _style_sheet(ws)

        # Drawdown check sheet.
        drawdown_df = _drawdown_check(conn)
        ws = wb.create_sheet("Drawdown_Check")
        _write_dataframe(ws, drawdown_df)
        ws.cell(row=1, column=5, value="excel_running_peak")
        ws.cell(row=1, column=6, value="excel_drawdown_formula")
        ws.cell(row=1, column=7, value="diff_vs_python_sql")
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=5, value=f"=MAX($C$2:C{row})")
            ws.cell(row=row, column=6, value=f"=C{row}/E{row}-1")
            ws.cell(row=row, column=7, value=f"=F{row}-D{row}")
        _style_sheet(ws)

        # Portfolio contribution check sheet.
        portfolio_df = _portfolio_check(conn)
        ws = wb.create_sheet("Portfolio_Check")
        _write_dataframe(ws, portfolio_df)
        ws.cell(row=1, column=6, value="excel_contribution")
        ws.cell(row=1, column=7, value="diff_vs_python_sql")
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=6, value=f"=C{row}*D{row}")
            ws.cell(row=row, column=7, value=f"=F{row}-E{row}")
        _style_sheet(ws)

        # Benchmark check sheet.
        benchmark_df = _benchmark_check(conn)
        ws = wb.create_sheet("Benchmark_Check")
        _write_dataframe(ws, benchmark_df)
        ws.cell(row=1, column=7, value="excel_active_return")
        ws.cell(row=1, column=8, value="diff_vs_python_sql")
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=7, value=f"=B{row}-C{row}")
            ws.cell(row=row, column=8, value=f"=G{row}-D{row}")
        _style_sheet(ws)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            if ws.title != "README":
                for cell in ws[1]:
                    cell.fill = YELLOW_FILL if "diff" in str(cell.value).lower() else HEADER_FILL

        wb.save(output_path)

    print(f"Step 15 complete: wrote validation workbook to {output_path}")


if __name__ == "__main__":
    create_excel_validation_workbook()
